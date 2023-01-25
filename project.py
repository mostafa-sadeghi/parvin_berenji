#  pip install openpyxl
#  pip install xlrd
#  pip install pillow
#  pip install matplotlib
#  pip install ttkbootstrap

from tkinter import ttk, END
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import ttkbootstrap as ttkb
import os
import openpyxl
import xlrd
from openpyxl import Workbook
import pathlib
from matplotlib import pyplot as plt


def clear():
    name_entry.delete(0, END)
    family_entry.delete(0, END)
    Postalcode_entry.delete(0, END)
    work_experience_entry.delete(0, END)
    salary_entry.delete(0, END)


def save():
    name_val = name_entry.get()
    family_val = family_entry.get()
    postal_code_val = Postalcode_entry.get()
    work_experience_val = work_experience_entry.get()
    salary_val = salary_entry.get()
    gender_val = radio.get()
    # write data on excel file:
    file = openpyxl.load_workbook('lab.xlsx')
    sheet = file.active
    sheet.cell(row=sheet.max_row+1, column=1, value=sheet.max_row)
    sheet.cell(row=sheet.max_row, column=2, value=name_val)
    sheet.cell(row=sheet.max_row, column=3, value=family_val)
    sheet.cell(row=sheet.max_row, column=4, value=postal_code_val)
    sheet.cell(row=sheet.max_row, column=5, value=work_experience_val)
    sheet.cell(row=sheet.max_row, column=6, value=salary_val)
    sheet.cell(row=sheet.max_row, column=7, value=gender_val)
    file.save('lab.xlsx')
    clear()


def plot(type):

    file = openpyxl.load_workbook('lab.xlsx')
    sheet = file.active
    salary = []
    experience = []
    gender = []

    for index, row in enumerate(sheet.rows):
        salary.append(sheet.cell(index+1, 6).value)
        experience.append(sheet.cell(index+1, 5).value)
        gender.append(sheet.cell(index+1, 7).value)

    salary = salary[1:]
    experience = experience[1:]
    gender = gender[1:]
    salary = [float(s) for s in salary]
    salary.sort()
    experience = [float(e) for e in experience]
    experience.sort()

    if type == 'plot':
        # plotting
        plt.plot(experience, salary, label="experience")
        # plt.plot(gender, salary, label="gender")
    elif type == 'bar':
        plt.bar(experience, salary, label="salary estimation")

    # plt.style.use('ggplot')
    plt.xkcd()
    plt.title('Salary Per Experience')
    plt.xlabel('Experience')
    plt.ylabel('Salary')
    plt.legend()
    # plt.grid()
    plt.tight_layout()
    plt.show()


root = tk.Tk()
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)
root.rowconfigure(1, weight=1)

style = ttkb.Style('superhero')
style.configure("TButton", font=("Helvetica", 12, "bold"))
root.title('Our Registration App')
root.iconbitmap('orange.ico')
# root.geometry("1250x700+110+30")
# root.resizable(False,False)
# root.config(bg=background)
file = pathlib.Path('lab.xlsx')

if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "ID"
    sheet['B1'] = "Name"
    sheet['C1'] = "Family"
    sheet['D1'] = "Postalcode"
    sheet['E1'] = "work experience"
    sheet['F1'] = "salary"
    sheet['G1'] = "gender"

    file.save('lab.xlsx')

input_frame = ttkb.Frame()
input_frame.grid(row=0, column=0)

name_label = ttkb.Label(
    input_frame, text="Name:", bootstyle="success", font=("Helvetica", 12, "bold"))
name_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")
name_entry = ttkb.Entry(input_frame, width=25)
name_entry.grid(row=0, column=1, padx=10, pady=10)

family_label = ttkb.Label(
    input_frame, text="Family:", bootstyle="success", font=("Helvetica", 12, "bold"))
family_label.grid(row=1, column=0, padx=10, pady=10, sticky="w")
family_entry = ttkb.Entry(input_frame, width=25)
family_entry.grid(row=1, column=1, padx=10, pady=10)

Postalcode_label = ttkb.Label(
    input_frame, text="Postalcode:", bootstyle="success", font=("Helvetica", 12, "bold"))
Postalcode_label.grid(row=2, column=0, padx=10, pady=10, sticky="w")
Postalcode_entry = ttkb.Entry(input_frame, width=25)
Postalcode_entry.grid(row=2, column=1, padx=10, pady=10)

work_experience_label = ttkb.Label(
    input_frame, text="work experience:", bootstyle="success", font=("Helvetica", 12, "bold"))
work_experience_label.grid(row=3, column=0, padx=10, pady=10, sticky="w")
work_experience_entry = ttkb.Entry(input_frame, width=25)
work_experience_entry.grid(row=3, column=1, padx=10, pady=10)

salary_label = ttkb.Label(
    input_frame, text="salary:", bootstyle="success", font=("Helvetica", 12, "bold"))
salary_label.grid(row=4, column=0, padx=10, pady=10, sticky="w")
salary_entry = ttkb.Entry(input_frame, width=25)
salary_entry.grid(row=4, column=1, padx=10, pady=10)

gender_label = ttkb.Label(input_frame, text="Gender",
                          bootstyle="success", font=("Helvetica", 12, "bold"))
gender_label.grid(row=5, column=0, padx=10, pady=10, sticky="w")
radio = tk.StringVar()

radio_frame = ttkb.Frame(input_frame)
radio_frame.grid(row=5, column=1, padx=10, pady=10)

male_radio = ttkb.Radiobutton(
    radio_frame, text="Male", variable=radio, value="Male")
male_radio.grid(row=0, column=0, padx=10, pady=10)
female_radio = ttkb.Radiobutton(
    radio_frame, text="Female", variable=radio, value="Female")
female_radio.grid(row=0, column=1, padx=10, pady=10)


# buttons frame

button_frame = ttkb.Frame(root)
button_frame.grid(row=1, column=0, pady=10)

save_button = ttkb.Button(button_frame, text="save",
                          bootstyle="success outline", command=save)
save_button.grid(row=0, column=0, sticky="ew", padx=10)
plot_button = ttkb.Button(button_frame, text="plot",
                          bootstyle="danger", command=lambda: plot('plot'))
plot_button.grid(row=0, column=1, sticky="ew", padx=10)
bar_plot_button = ttkb.Button(button_frame, text="barplot",
                              bootstyle="primary", command=lambda: plot('bar'))
bar_plot_button.grid(row=0, column=2, sticky="ew", padx=10)
quit_button = ttkb.Button(button_frame, text="quit",
                          bootstyle="secondary", command=root.destroy)
quit_button.grid(row=0, column=3, sticky="ew", padx=10)

root.mainloop()
